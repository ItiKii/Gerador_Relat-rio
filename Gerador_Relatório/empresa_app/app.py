from flask import Flask, render_template, request, send_file
from openpyxl import Workbook
from openpyxl.chart import PieChart, BarChart, Reference, LineChart
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, NamedStyle
from openpyxl.formatting.rule import ColorScaleRule
from io import BytesIO
from datetime import datetime

app = Flask(__name__)


def converter_moeda(valor_str):
    """Converte moeda brasileira para float"""
    return float(valor_str.replace(".", "").replace(",", "."))


def formatar_moeda(valor):
    """Formata valor para exibição"""
    return f"R$ {valor:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")


def criar_planilha(dados):
    wb = Workbook()
    ws = wb.active
    ws.title = "Relatório Mensal"
    
    # Configurar larguras das colunas
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 20
    ws.column_dimensions["F"].width = 15
    
    # Cores personalizadas
    cor_titulo = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    cor_subtitulo = PatternFill(start_color="34495E", end_color="34495E", fill_type="solid")
    cor_destaque = PatternFill(start_color="27AE60", end_color="27AE60", fill_type="solid")
    cor_alerta = PatternFill(start_color="E74C3C", end_color="E74C3C", fill_type="solid")
    cor_info = PatternFill(start_color="3498DB", end_color="3498DB", fill_type="solid")
    
    # Estilos de borda
    thin_border = Border(
        left=Side(style='thin'), 
        right=Side(style='thin'), 
        top=Side(style='thin'), 
        bottom=Side(style='thin')
    )
    
    # ======================
    # TÍTULO PRINCIPAL
    # ======================
    ws.merge_cells("A1:F1")
    titulo = ws["A1"]
    titulo.value = "📊 RELATÓRIO DE DESEMPENHO COMERCIAL"
    titulo.font = Font(size=18, bold=True, color="FFFFFF")
    titulo.alignment = Alignment(horizontal="center", vertical="center")
    titulo.fill = cor_titulo
    ws.row_dimensions[1].height = 35
    
    # ======================
    # INFORMAÇÕES GERAIS
    # ======================
    ws.merge_cells("A2:F2")
    subtitulo = ws["A2"]
    mes_ano = dados["mes"]
    data_geracao = datetime.now().strftime("%d/%m/%Y %H:%M")
    subtitulo.value = f"Período: {mes_ano} | Produto: {dados['item']} | Gerado em: {data_geracao}"
    subtitulo.font = Font(size=11, italic=True, color="FFFFFF")
    subtitulo.alignment = Alignment(horizontal="center")
    subtitulo.fill = cor_subtitulo
    
    ws.append([])
    
    # ======================
    # INDICADORES RESUMO (KPI Cards)
    # ======================
    ws.merge_cells("A4:C4")
    ws["A4"] = "📈 INDICADORES PRINCIPAIS"
    ws["A4"].font = Font(size=14, bold=True, color="2C3E50")
    ws["A4"].fill = PatternFill(start_color="F2F4F4", end_color="F2F4F4", fill_type="solid")
    ws["A4"].border = thin_border
    
    # Calcular percentuais
    margem_lucro_percentual = (dados["lucro_total"] / dados["receita_total"] * 100) if dados["receita_total"] > 0 else 0
    ocupacao_estoque = (dados["vendidos"] / (dados["vendidos"] + dados["estoque_restante"]) * 100) if (dados["vendidos"] + dados["estoque_restante"]) > 0 else 0
    
    kpis = [
        ("📦 Receita Total", formatar_moeda(dados["receita_total"]), "cor_info"),
        ("💰 Lucro Total", formatar_moeda(dados["lucro_total"]), "cor_destaque"),
        ("📊 Margem de Lucro", f"{margem_lucro_percentual:.1f}%", "cor_info"),
        ("🏭 Custo Total", formatar_moeda(dados["custo_total"]), "cor_alerta"),
        ("📋 Unidades Vendidas", f"{dados['vendidos']:,}", "cor_info"),
        ("📦 Estoque Restante", f"{dados['estoque_restante']:,} ({ocupacao_estoque:.0f}% vendido)", "cor_info"),
    ]
    
    linha_atual = 5
    for i in range(0, len(kpis), 2):
        for j in range(2):
            if i + j < len(kpis):
                col = chr(65 + (j * 3))  # Coluna A ou D
                ws.merge_cells(f"{col}{linha_atual}:{chr(65 + (j * 3) + 2)}{linha_atual}")
                cell = ws[f"{col}{linha_atual}"]
                cell.value = f"{kpis[i+j][0]}: {kpis[i+j][1]}"
                cell.font = Font(size=12, bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border
                
                # Aplicar cor baseada no tipo
                if "lucro" in kpis[i+j][0].lower() or "margem" in kpis[i+j][0].lower():
                    cell.fill = cor_destaque
                    cell.font = Font(size=12, bold=True, color="FFFFFF")
                elif "custo" in kpis[i+j][0].lower():
                    cell.fill = cor_alerta
                    cell.font = Font(size=12, bold=True, color="FFFFFF")
                else:
                    cell.fill = cor_info
                    cell.font = Font(size=12, bold=True, color="FFFFFF")
        linha_atual += 1
    
    linha_atual += 1
    
    # ======================
    # TABELA DE ANÁLISE DETALHADA
    # ======================
    ws.merge_cells(f"A{linha_atual}:F{linha_atual}")
    ws[f"A{linha_atual}"] = "📋 ANÁLISE DETALHADA"
    ws[f"A{linha_atual}"].font = Font(size=14, bold=True, color="2C3E50")
    ws[f"A{linha_atual}"].fill = PatternFill(start_color="F2F4F4", end_color="F2F4F4", fill_type="solid")
    ws[f"A{linha_atual}"].border = thin_border
    linha_atual += 1
    
    # Cabeçalho da tabela
    cabecalhos = ["Métrica", "Valor", "Unidade", "Percentual", "Status", "Análise"]
    for i, cab in enumerate(cabecalhos):
        col = chr(65 + i)
        cell = ws[f"{col}{linha_atual}"]
        cell.value = cab
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = cor_titulo
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border
    
    linha_atual += 1
    
    # Dados detalhados
    dados_detalhados = [
        ["Valor Unitário", dados["valor_item"], "R$", "100%", "Base", "Preço de venda"],
        ["Custo de Produção", dados["custo_produto"], "R$", f"{(dados['custo_produto']/dados['valor_item']*100):.1f}%", "Custo", "Por unidade"],
        ["Lucro Unitário", dados["lucro_unitario"], "R$", f"{(dados['lucro_unitario']/dados['valor_item']*100):.1f}%", "Lucro", "Margem por unidade"],
        ["Quantidade Vendida", dados["vendidos"], "un", f"{(dados['vendidos']/(dados['vendidos']+dados['estoque_restante'])*100):.0f}%", "Vendas", "Do estoque total"],
        ["Receita Total", dados["receita_total"], "R$", "100%", "Receita", "Faturamento bruto"],
        ["Custo Total", dados["custo_total"], "R$", f"{(dados['custo_total']/dados['receita_total']*100):.1f}%", "Custo", "Custo de produção total"],
        ["Lucro Total", dados["lucro_total"], "R$", f"{margem_lucro_percentual:.1f}%", "Lucro", "Resultado final"],
        ["Estoque Restante", dados["estoque_restante"], "un", f"{(100-ocupacao_estoque):.0f}%", "Estoque", f"Disponível para venda"],
    ]
    
    for i, dados_linha in enumerate(dados_detalhados):
        for j, valor in enumerate(dados_linha):
            col = chr(65 + j)
            cell = ws[f"{col}{linha_atual + i}"]
            cell.value = valor
            cell.alignment = Alignment(horizontal="center" if j > 1 else "left")
            cell.border = thin_border
            
            # Formatação de moeda para valores monetários
            if j == 1 and dados_linha[2] == "R$":
                cell.number_format = 'R$ #,##0.00'
            
            # Cores condicionais baseadas no status
            if j == 4:  # Coluna Status
                if "Lucro" in valor:
                    cell.fill = PatternFill(start_color="27AE60", end_color="27AE60", fill_type="solid")
                    cell.font = Font(bold=True, color="FFFFFF")
                elif "Custo" in valor:
                    cell.fill = PatternFill(start_color="E74C3C", end_color="E74C3C", fill_type="solid")
                    cell.font = Font(bold=True, color="FFFFFF")
                elif "Receita" in valor:
                    cell.fill = PatternFill(start_color="3498DB", end_color="3498DB", fill_type="solid")
                    cell.font = Font(bold=True, color="FFFFFF")
                elif "Vendas" in valor:
                    cell.fill = PatternFill(start_color="F39C12", end_color="F39C12", fill_type="solid")
                    cell.font = Font(bold=True, color="FFFFFF")
    
    linha_atual += len(dados_detalhados) + 2
    
    # ======================
    # GRÁFICOS INFORMATIVOS
    # ======================
    # Título dos gráficos
    ws.merge_cells(f"A{linha_atual}:F{linha_atual}")
    ws[f"A{linha_atual}"] = "📊 ANÁLISE GRÁFICA"
    ws[f"A{linha_atual}"].font = Font(size=14, bold=True, color="2C3E50")
    ws[f"A{linha_atual}"].fill = PatternFill(start_color="F2F4F4", end_color="F2F4F4", fill_type="solid")
    ws[f"A{linha_atual}"].border = thin_border
    linha_atual += 2
    
    # Criar dados para gráficos
    dados_grafico = [
        ["Categoria", "Valor (R$)", "Percentual"],
        ["Custo Total", dados["custo_total"], f"{(dados['custo_total']/dados['receita_total']*100):.1f}%"],
        ["Lucro Total", dados["lucro_total"], f"{margem_lucro_percentual:.1f}%"],
    ]
    
    inicio_grafico = linha_atual
    for dados_linha in dados_grafico:
        ws.append(dados_linha)
        for col in range(1, 4):
            cell = ws.cell(row=linha_atual, column=col)
            cell.border = thin_border
            if linha_atual == inicio_grafico:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = cor_titulo
        linha_atual += 1
    
    # GRÁFICO 1: Gráfico de Pizza (Distribuição Custo vs Lucro)
    pie = PieChart()
    pie.title = "Distribuição Custo vs Lucro"
    pie.height = 10
    pie.width = 15
    
    labels = Reference(ws, min_col=1, min_row=inicio_grafico + 1, max_row=inicio_grafico + 2)
    data = Reference(ws, min_col=2, min_row=inicio_grafico, max_row=inicio_grafico + 2)
    
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(labels)
    
    # Personalizar cores do gráfico de pizza
    pie.series[0].graphicalProperties.solidFill = "3498DB"  # Azul para custo
    if len(pie.series[0].data_points) > 1:
        pie.series[0].data_points[0].graphicalProperties.solidFill = "E74C3C"  # Vermelho para custo
        pie.series[0].data_points[1].graphicalProperties.solidFill = "27AE60"  # Verde para lucro
    
    # GRÁFICO 2: Gráfico de Barras (Comparativo)
    bar_chart = BarChart()
    bar_chart.title = "Análise Comparativa"
    bar_chart.height = 10
    bar_chart.width = 20
    bar_chart.y_axis.title = "Valor (R$)"
    bar_chart.x_axis.title = "Categorias"
    
    # Dados para gráfico de barras
    ws.append([])
    ws.append(["Indicador", "Valor"])
    ws.append(["Receita", dados["receita_total"]])
    ws.append(["Custo", dados["custo_total"]])
    ws.append(["Lucro", dados["lucro_total"]])
    
    data_bar = Reference(ws, min_col=2, min_row=linha_atual + 1, max_row=linha_atual + 3)
    categories_bar = Reference(ws, min_col=1, min_row=linha_atual + 1, max_row=linha_atual + 3)
    
    bar_chart.add_data(data_bar, titles_from_data=True)
    bar_chart.set_categories(categories_bar)
    
    # Adicionar gráficos à planilha
    ws.add_chart(pie, f"A{linha_atual + 5}")
    ws.add_chart(bar_chart, f"D{linha_atual + 5}")
    
    # ======================
    # ANÁLISE DE TENDÊNCIAS (simulada)
    # ======================
    linha_atual += 15
    
    ws.merge_cells(f"A{linha_atual}:F{linha_atual}")
    ws[f"A{linha_atual}"] = "📈 ANÁLISE DE TENDÊNCIAS E RECOMENDAÇÕES"
    ws[f"A{linha_atual}"].font = Font(size=14, bold=True, color="2C3E50")
    ws[f"A{linha_atual}"].fill = PatternFill(start_color="F2F4F4", end_color="F2F4F4", fill_type="solid")
    ws[f"A{linha_atual}"].border = thin_border
    linha_atual += 1
    
    # Análise de performance
    if margem_lucro_percentual > 30:
        analise_lucro = "Excelente margem de lucro! ⭐"
        cor_analise = "27AE60"
    elif margem_lucro_percentual > 15:
        analise_lucro = "Margem de lucro satisfatória ✓"
        cor_analise = "F39C12"
    else:
        analise_lucro = "Margem de lucro baixa, rever custos ⚠"
        cor_analise = "E74C3C"
    
    if ocupacao_estoque > 70:
        analise_estoque = "Boa rotatividade de estoque ✓"
    else:
        analise_estoque = "Estoque com baixa rotatividade, considerar promoções"
    
    recomendacoes = [
        ["Análise de Rentabilidade:", f"Margem Líquida: {margem_lucro_percentual:.1f}%", analise_lucro],
        ["Gestão de Estoque:", f"Taxa de Ocupação: {ocupacao_estoque:.0f}%", analise_estoque],
        ["Recomendação 1:", "Manter preço atual", "Considerar aumento se mercado permitir"],
        ["Recomendação 2:", "Otimizar custos", f"Reduzir custos em 5% aumentaria lucro em {((dados['lucro_total'] + dados['custo_total']*0.05)/dados['lucro_total']*100 - 100):.0f}%"],
        ["Projeção:", f"Lucro projetado (próximo mês)", formatar_moeda(dados["lucro_total"] * 1.1)],
    ]
    
    for i, rec in enumerate(recomendacoes):
        for j, texto in enumerate(rec):
            col = chr(65 + j*2)
            ws.merge_cells(f"{col}{linha_atual + i}:{chr(65 + j*2 + 1)}{linha_atual + i}")
            cell = ws[f"{col}{linha_atual + i}"]
            cell.value = texto
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="left")
            
            if j == 0:
                cell.font = Font(bold=True)
            if j == 2 and "⭐" in texto:
                cell.font = Font(color="27AE60", bold=True)
            elif j == 2 and "⚠" in texto:
                cell.font = Font(color="E74C3C", bold=True)
    
    # ======================
    # RODAPÉ
    # ======================
    linha_atual += 7
    ws.merge_cells(f"A{linha_atual}:F{linha_atual}")
    rodape = ws[f"A{linha_atual}"]
    rodape.value = "Relatório gerado automaticamente pelo sistema de gestão - Todos os direitos reservados"
    rodape.font = Font(size=9, italic=True, color="7F8C8D")
    rodape.alignment = Alignment(horizontal="center")
    
    # Formatação condicional para valores
    ws.conditional_formatting.add(f'B5:B20', ColorScaleRule(
        start_type='min', start_color='FFEEEE',
        mid_type='percentile', mid_value=50, mid_color='FFFFFF',
        end_type='max', end_color='EEFFEE'
    ))
    
    return wb


@app.route("/", methods=["GET", "POST"])
def index():
    if request.method == "POST":
        valor_item = converter_moeda(request.form["valor_item"])
        custo_produto = converter_moeda(request.form["custo_produto"])

        vendidos = int(request.form["vendidos"])
        estoque = int(request.form["estoque"])

        dados = {
            "item": request.form["item"],
            "mes": request.form["mes"],
            "valor_item": valor_item,
            "custo_produto": custo_produto,
            "vendidos": vendidos,
            "lucro_unitario": valor_item - custo_produto,
            "receita_total": valor_item * vendidos,
            "custo_total": custo_produto * vendidos,
            "lucro_total": (valor_item * vendidos) - (custo_produto * vendidos),
            "estoque_restante": estoque - vendidos
        }

        wb = criar_planilha(dados)

        arquivo = BytesIO()
        wb.save(arquivo)
        arquivo.seek(0)

        return send_file(
            arquivo,
            as_attachment=True,
            download_name=f"relatorio_{dados['item']}_{dados['mes']}.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    return render_template("index.html")


if __name__ == "__main__":
    app.run(debug=True)